import os
import sys
import cv2
import glob
import random
import itertools
import numpy as np
import tensorflow as tf
from scipy.io import loadmat
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill
from tensorflow.python.keras import backend as k


# get_img_seg & data_loader give input data and label
def get_img_seg(path_img, path_softmax, path_seg, height, width, num_classes, resize):
    img = cv2.imread(path_img)
    softmax = np.load(path_softmax)
    seg = cv2.imread(path_seg, cv2.IMREAD_GRAYSCALE)

    img = img / 127.5 - 1

    h = img.shape[0]
    w = img.shape[1]

    # each layer of this array is a mask for a specific object
    if resize:

        if h <= w:

            start = random.randint(0, w - h)

            img = img[0:h, start: start + h]
            img = cv2.resize(src=img, dsize=(height, width), interpolation=cv2.INTER_LINEAR)

            softmax = softmax[0:h, start: start + h]
            softmax = cv2.resize(src=softmax, dsize=(height, width), interpolation=cv2.INTER_LINEAR)

            seg = seg[0:h, start: start + h]
            seg = cv2.resize(src=seg, dsize=(height, width), interpolation=cv2.INTER_NEAREST)

        else:

            start = random.randint(0, h - w)

            img = img[start:start + w, 0: w]
            img = cv2.resize(src=img, dsize=(height, width), interpolation=cv2.INTER_LINEAR)

            softmax = softmax[start:start + w, 0: w]
            softmax = cv2.resize(src=softmax, dsize=(height, width), interpolation=cv2.INTER_LINEAR)

            seg = seg[start:start + w, 0: w]
            seg = cv2.resize(src=seg, dsize=(height, width), interpolation=cv2.INTER_NEAREST)

    seg_labels = tf.keras.utils.to_categorical(y=seg, num_classes=256, dtype='uint8')
    seg_labels = seg_labels[:, :, 0:num_classes]

    return img, softmax, seg_labels


def data_loader(dir_img, dir_seg, dir_softmax, batch_size, h, w, num_classes, resize):
    # list of all image path png
    print(dir_img)
    images = glob.glob(dir_img + "*.png")
    images.sort()

    print(dir_softmax)
    images_softmax = glob.glob(dir_softmax + "*.npy")
    images_softmax.sort()

    # list of all seg img path
    print(dir_seg)
    segmentations = glob.glob(dir_seg + "*.png")
    segmentations.sort()

    # create an iterator of tuples ( img and its seg_img)
    zipped = itertools.cycle(zip(images, images_softmax, segmentations))

    while 1:

        X = []
        S = []
        Y = []

        for _ in range(batch_size):
            img_path, softmax_path, seg_path = next(zipped)
            i, sf, s = get_img_seg(path_img=img_path, path_softmax=softmax_path, path_seg=seg_path, height=h, width=w,
                                   num_classes=num_classes,
                                   resize=resize)

            X.append(i)
            S.append(sf)
            Y.append(s)

        yield [np.array(X), np.array(S)], np.array(Y)


def get_img_seg_baseline(path_img, path_seg, height, width, num_classes, resize):
    img = cv2.imread(path_img)
    img = img / 127.5 - 1
    seg = cv2.imread(path_seg, cv2.IMREAD_GRAYSCALE)

    h = img.shape[0]
    w = img.shape[1]

    # each layer of this array is a mask for a specific object
    if resize:

        # seg_labels = np.zeros((height, width, num_classes))

        if h <= w:

            start = random.randint(0, w - h)

            img = img[0:h, start: start + h]
            img = cv2.resize(src=img, dsize=(height, width), interpolation=cv2.INTER_LINEAR)

            seg = seg[0:h, start: start + h]
            seg = cv2.resize(src=seg, dsize=(height, width), interpolation=cv2.INTER_NEAREST)

        else:

            start = random.randint(0, h - w)

            img = img[start:start + w, 0: w]
            img = cv2.resize(src=img, dsize=(height, width), interpolation=cv2.INTER_LINEAR)

            seg = seg[start:start + w, 0: w]
            seg = cv2.resize(src=seg, dsize=(height, width), interpolation=cv2.INTER_NEAREST)

    seg_labels = tf.keras.utils.to_categorical(y=seg, num_classes=256, dtype='uint8')
    seg_labels = seg_labels[:, :, 0:num_classes]

    return img, seg_labels


def data_loader_baseline(dir_img, dir_seg, batch_size, h, w, num_classes, resize):
    # list of all image path png
    print(dir_img)
    images = glob.glob(dir_img + "*.png")
    images.sort()
    # list of all seg img path
    print(dir_seg)
    segmentations = glob.glob(dir_seg + "*.png")
    segmentations.sort()

    # create an iterator of tuples ( img and its seg_img)
    zipped = itertools.cycle(zip(images, segmentations))

    while 1:

        X = []
        Y = []

        for _ in range(batch_size):
            im_path, seg_path = next(zipped)
            i, s = get_img_seg_baseline(im_path, seg_path, h, w, num_classes, resize)
            X.append(i)
            Y.append(s)

        yield np.array(X), np.array(Y)


def calc_adj_mat(batch_imgs, batch_size):
    adj_mat = k.zeros(shape=(108, 108))

    for o in range(batch_size):
        img = batch_imgs[o]
        classes = np.unique(img)
        classes = classes[1:]
        if 255 in classes:
            classes = classes[:-1]
        mat_contour = []

        for i in range(len(classes)):

            value = classes[i]
            mask = cv2.inRange(img, int(value), int(value))
            per, _ = cv2.findContours(image=mask, mode=cv2.RETR_TREE, method=cv2.CHAIN_APPROX_SIMPLE)

            mat_total = k.zeros(shape=(1, 2))

            for q in range(len(per)):

                tmp = per[q]
                mat = k.zeros(shape=(len(tmp), 2))
                for j in range(len(tmp)):
                    point = tmp[j]
                    x = point[0][0]
                    y = point[0][1]
                    mat[j][0] = x
                    mat[j][1] = y

                mat_total = k.concatenate((mat_total, mat), axis=0)

            mat_contour.append(mat_total[1:])

        for i in range(len(classes)):
            tmp = mat_contour[i]

            for j in range(i + 1, len(classes)):
                # for j in range(0, len(classes)):
                min_v = sys.maxsize
                second_mat = mat_contour[j]

                for p in range(len(tmp)):
                    first_mat = tmp[p]

                    dif = first_mat - second_mat
                    # dif = np.multiply(dif, dif)
                    dif = dif * dif
                    sum_mat = k.sum(dif, 1)
                    sqrt = k.sqrt(sum_mat)

                    min_tmp = k.min(sqrt)
                    if min_tmp < min_v:
                        min_v = min_tmp

                if min_v <= 1:
                    adj_mat[classes[i]][classes[j]] = 1 + adj_mat[classes[i]][classes[j]]

    # adj_mat = normalize(adj_mat, axis=1, norm='l1')

    return adj_mat


def calc_adj_mat_error(batch_imgs, batch_size):
    adj_mat = k.zeros(shape=(108, 108))

    for o in range(batch_size):
        img = batch_imgs[o]
        classes = np.unique(img)
        classes = classes[1:]
        if 255 in classes:
            classes = classes[:-1]
        mat_contour = []

        for i in range(len(classes)):

            value = classes[i]
            mask = cv2.inRange(img, int(value), int(value))
            per, _ = cv2.findContours(image=mask, mode=cv2.RETR_TREE, method=cv2.CHAIN_APPROX_SIMPLE)

            mat_total = k.zeros(shape=(1, 2))

            for q in range(len(per)):

                tmp = per[q]
                mat = k.zeros(shape=(len(tmp), 2))
                for j in range(len(tmp)):
                    point = tmp[j]
                    x = point[0][0]
                    y = point[0][1]
                    mat[j][0] = x
                    mat[j][1] = y

                mat_total = k.concatenate((mat_total, mat), axis=0)

            mat_contour.append(mat_total[1:])

        for i in range(len(classes)):
            tmp = mat_contour[i]

            for j in range(i + 1, len(classes)):
                # for j in range(0, len(classes)):
                min_v = sys.maxsize
                second_mat = mat_contour[j]

                for p in range(len(tmp)):
                    first_mat = tmp[p]

                    dif = first_mat - second_mat
                    # dif = np.multiply(dif, dif)
                    dif = dif * dif
                    sum_mat = k.sum(dif, 1)
                    sqrt = k.sqrt(sum_mat)

                    min_tmp = k.min(sqrt)
                    if min_tmp < min_v:
                        min_v = min_tmp

                if min_v <= 1:
                    adj_mat[classes[i]][classes[j]] = 1 + adj_mat[classes[i]][classes[j]]

    # adj_mat = normalize(adj_mat, axis=1, norm='l1')

    return adj_mat


def mapCl2Prt():
    mapPart2Classes = [
        [0, 1],
        [1, 6],
        [6, 10],
        [10, 18],
        [18, 19],
        [19, 21],
        [21, 29],
        [29, 36],
        [36, 45],
        [45, 46],
        [46, 54],
        [54, 55],
        [55, 65],
        [65, 73],
        [73, 76],
        [77, 89],
        [89, 90],
        [91, 98],
        [99, 100],
        [100, 107],
        [107, 108],
    ]
    return mapPart2Classes


def listPartsNames():
    listParts = ['background', 'aeroplane_body', 'aeroplane_stern', 'aeroplane_rwing',
                 'aeroplane_engine', 'aeroplane_wheel',
                 'bicycle_fwheel', 'bicycle_saddle', 'bicycle_handlebar', 'bicycle_chainwheel',
                 'birds_head', 'birds_beak',
                 'birds_torso', 'birds_neck', 'birds_rwing', 'birds_rleg', 'birds_rfoot',
                 'birds_tail', 'boat', 'bottle_cap',
                 'bottle_body', 'bus_rightside', 'bus_roofside', 'bus_rightmirror', 'bus_fliplate',
                 'bus_door',
                 'bus_wheel', 'bus_headlight', 'bus_window', 'car_rightside', 'car_roofside',
                 'car_fliplate',
                 'car_door', 'car_wheel', 'car_headlight', 'car_window', 'cat_head', 'cat_reye',
                 'cat_rear',
                 'cat_nose', 'cat_torso', 'cat_neck', 'cat_rfleg', 'cat_rfpa', 'cat_tail', 'chair',
                 'cow_head', 'cow_rear',
                 'cow_muzzle', 'cow_rhorn', 'cow_torso', 'cow_neck', 'cow_rfuleg', 'cow_tail',
                 'diningtable', 'dog_head',
                 'dog_reye', 'dog_rear', 'dog_nose', 'dog_torso', 'dog_neck', 'dog_rfleg',
                 'dog_rfpa', 'dog_tail',
                 'dog_muzzle', 'horse_head', 'horse_rear', 'horse_muzzle', 'horse_torso',
                 'horse_neck', 'horse_rfuleg',
                 'horse_tail', 'horse_rfho', 'motorbike_fwheel', 'motorbike_handlebar',
                 'motorbike_saddle',
                 'motorbike_headlight', 'person_head', 'person_reye', 'person_rear', 'person_nose',
                 'person_mouth',
                 'person_hair', 'person_torso', 'person_neck', 'person_ruarm', 'person_rhand',
                 'person_ruleg',
                 'person_rfoot', 'pottedplant_pot', 'pottedplant_plant', 'sheep_head', 'sheep_rear',
                 'sheep_muzzle',
                 'sheep_rhorn', 'sheep_torso', 'sheep_neck', 'sheep_rfuleg', 'sheep_tail', 'sofa',
                 'train_head',
                 'train_hrightside', 'train_hroofside', 'train_headlight', 'train_coach',
                 'train_crightside',
                 'train_croofside', 'tvmonitor_screen']

    return listParts


def dictImages():
    img_dict = {
        "2008_000045.png": "Treno",
        "2008_000093.png": "Divano",
        "2008_000142.png": "Persona e cavallo",
        "2008_000689.png": "Moto",
        "2008_000585.png": "Aereo",
        "2008_001047.png": "Barca",
        "2008_001704.png": "Schermo",
        "2008_001770.png": "Uccello",
        "2008_002062.png": "Macchina",
        "2008_002583.png": "Gatto",
        "2008_001434.png": "Tavolo"
    }
    return img_dict


def createDirectories(prefix, lr_p, batch_sz, h_img, mult_rate, dil_rate, use_BN):
    path = "./" + prefix + "_class_108_lr_" + str(lr_p) + "_batch_" + str(
        batch_sz) + "_size_" + str(h_img)

    if dil_rate:
        path = path + "_use_dil_rate"
    if mult_rate > 1:
        path = path + "_use_mult_rate_" + str(mult_rate) + ""
    if use_BN:
        path = path + "_use_BN"

    path = path + "/"

    # print(path)

    if not os.path.isdir(path):
        os.mkdir(path)

    pathTBoard = "./" + path + "Graph_deeplab/"
    if not os.path.isdir(pathTBoard):
        os.mkdir(pathTBoard)

    pT = pathTBoard + datetime.now().strftime("%Y%m%d-%H%M%S")
    if not os.path.isdir(pT):
        os.mkdir(pT)

    pathTChPoints = "./" + path + "Checkpoints_deeplab/"
    if not os.path.isdir(pathTChPoints):
        os.mkdir(pathTChPoints)

    pathWeight = "./" + path + "Weight_deeplab/"
    if not os.path.isdir(pathWeight):
        os.mkdir(pathWeight)

    return path, pT, pathTChPoints, pathWeight


def list_mult_lr(factor):
    list = {'conv1_simple': factor,
            'conv1_BN_simple': factor,
            'conv2_simple': factor,
            'conv2_BN_simple': factor,
            'conv3_simple': factor,
            'conv3_BN_simple': factor,
            'conv4': factor,
            'conv4_BN_simple': factor,
            }
    return list


def print_var(num_classes, batch_sz, pathTr, pathTrSeg, pathVal, pathValSeg, h, w, tr_sz, val_sz):
    # Print var
    print('Variables')

    print('num classes: ' + str(num_classes))
    print('batch size: ' + str(batch_sz))
    print('img height: ' + str(h))
    print('img width: ' + str(w))
    print('path imgs train: ' + pathTr)
    print('path imgs train seg: ' + pathTrSeg)
    print('dt train size: ' + str(tr_sz))
    print('path imgs val: ' + pathVal)
    print('path imgs val seg: ' + pathValSeg)
    print('dt val size: ' + str(val_sz))


def listClassesNames():
    listParts = ['background',
                 'airplane',
                 'bicycle',
                 'bird',
                 'boat',
                 'bottle',
                 'bus',
                 'car',
                 'cat',
                 'chair',
                 'cow',
                 'table',
                 'dog',
                 'horse',
                 'motorbike',
                 'person',
                 'potted_plant',
                 'sheep',
                 'sofa',
                 'train',
                 'tv']

    return listParts


def create_excel_file(fileName="results", results21=None, results108=None, path=""):
    wb = Workbook()
    dest_filename = path + fileName + '.xlsx'
    ws1 = wb.active
    ws1.title = "results"

    pathCMap = 'Y:/tesisti/rossi/cmap255.mat'
    fileMat = loadmat(pathCMap)
    cmap = fileMat['cmap']

    # color map aRGB hex value
    map = []
    for i in range(len(cmap)):
        value = cmap[i]
        value0 = value[0]
        value1 = value[1]
        value2 = value[2]
        value = ('#{:02x}{:02x}{:02x}'.format(value0, value1, value2))
        map.append(value[1:])

    map_part = []
    map_part.append(1)
    map_part.append(2)
    map_part.append(7)
    map_part.append(11)
    map_part.append(19)
    map_part.append(20)
    map_part.append(22)
    map_part.append(30)
    map_part.append(37)
    map_part.append(46)
    map_part.append(47)
    map_part.append(55)
    map_part.append(56)
    map_part.append(66)
    map_part.append(74)
    map_part.append(78)
    map_part.append(90)
    map_part.append(92)
    map_part.append(100)
    map_part.append(101)
    map_part.append(108)

    ws1.merge_cells(start_row=1, end_row=1, end_column=1, start_column=1)
    ws1.merge_cells(start_row=2, end_row=6, end_column=1, start_column=1)
    ws1.merge_cells(start_row=7, end_row=10, end_column=1, start_column=1)
    ws1.merge_cells(start_row=11, end_row=18, end_column=1, start_column=1)
    ws1.merge_cells(start_row=19, end_row=19, end_column=1, start_column=1)
    ws1.merge_cells(start_row=20, end_row=21, end_column=1, start_column=1)
    ws1.merge_cells(start_row=22, end_row=29, end_column=1, start_column=1)
    ws1.merge_cells(start_row=30, end_row=36, end_column=1, start_column=1)
    ws1.merge_cells(start_row=37, end_row=45, end_column=1, start_column=1)
    ws1.merge_cells(start_row=46, end_row=46, end_column=1, start_column=1)
    ws1.merge_cells(start_row=47, end_row=54, end_column=1, start_column=1)
    ws1.merge_cells(start_row=55, end_row=55, end_column=1, start_column=1)
    ws1.merge_cells(start_row=56, end_row=65, end_column=1, start_column=1)
    ws1.merge_cells(start_row=66, end_row=73, end_column=1, start_column=1)
    ws1.merge_cells(start_row=74, end_row=77, end_column=1, start_column=1)
    ws1.merge_cells(start_row=78, end_row=89, end_column=1, start_column=1)
    ws1.merge_cells(start_row=90, end_row=91, end_column=1, start_column=1)
    ws1.merge_cells(start_row=92, end_row=99, end_column=1, start_column=1)
    ws1.merge_cells(start_row=100, end_row=100, end_column=1, start_column=1)
    ws1.merge_cells(start_row=101, end_row=107, end_column=1, start_column=1)
    ws1.merge_cells(start_row=108, end_row=108, end_column=1, start_column=1)

    ws1.merge_cells(start_row=1, end_row=1, end_column=2, start_column=2)
    ws1.merge_cells(start_row=2, end_row=6, end_column=2, start_column=2)
    ws1.merge_cells(start_row=7, end_row=10, end_column=2, start_column=2)
    ws1.merge_cells(start_row=11, end_row=18, end_column=2, start_column=2)
    ws1.merge_cells(start_row=19, end_row=19, end_column=2, start_column=2)
    ws1.merge_cells(start_row=20, end_row=21, end_column=2, start_column=2)
    ws1.merge_cells(start_row=22, end_row=29, end_column=2, start_column=2)
    ws1.merge_cells(start_row=30, end_row=36, end_column=2, start_column=2)
    ws1.merge_cells(start_row=37, end_row=45, end_column=2, start_column=2)
    ws1.merge_cells(start_row=46, end_row=46, end_column=2, start_column=2)
    ws1.merge_cells(start_row=47, end_row=54, end_column=2, start_column=2)
    ws1.merge_cells(start_row=55, end_row=55, end_column=2, start_column=2)
    ws1.merge_cells(start_row=56, end_row=65, end_column=2, start_column=2)
    ws1.merge_cells(start_row=66, end_row=73, end_column=2, start_column=2)
    ws1.merge_cells(start_row=74, end_row=77, end_column=2, start_column=2)
    ws1.merge_cells(start_row=78, end_row=89, end_column=2, start_column=2)
    ws1.merge_cells(start_row=90, end_row=91, end_column=2, start_column=2)
    ws1.merge_cells(start_row=92, end_row=99, end_column=2, start_column=2)
    ws1.merge_cells(start_row=100, end_row=100, end_column=2, start_column=2)
    ws1.merge_cells(start_row=101, end_row=107, end_column=2, start_column=2)
    ws1.merge_cells(start_row=108, end_row=108, end_column=2, start_column=2)

    classes = listClassesNames()
    index_class = 0
    for row in map_part:
        cell = ws1.cell(column=1, row=row, value="{0}".format(classes[index_class]))

        if results21 is not None:
            _ = ws1.cell(column=2, row=row, value="{0}".format(results21[index_class]))

        if index_class != 0:
            cell.fill = PatternFill("solid", fgColor=(map[index_class]))
        index_class = index_class + 1

    parts = listPartsNames()
    for row in range(len(parts)):
        cell = ws1.cell(column=3, row=row + 1, value="{0}".format(parts[row]))
        if results108 is not None:
            _ = ws1.cell(column=4, row=row + 1, value="{0}".format(results108[row]))
        if row != 0:
            cell.fill = PatternFill("solid", fgColor=(map[row]))

    wb.save(filename=dest_filename)
